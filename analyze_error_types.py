"""
analyze_error_types.py - 라벨링 시트 기반 유형별 교정률 분석

사용법:
  python analyze_error_types.py \
    --healed ./results/healed_labeling_29.xlsx \
    --failed ./results/error_labeling_233.xlsx
"""

import argparse
import re
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl이 필요합니다: pip install openpyxl")
    exit(1)


def extract_types(label):
    """라벨에서 개별 오류 유형 추출 (COMPLEX 내부 포함)"""
    label = str(label or '').strip()
    if not label:
        return []

    main = label.split('(')[0].strip() if '(' in label else label
    if main == 'COMPLEX':
        match = re.search(r'\(([^)]+)\)', label)
        if match:
            return [p.strip() for p in match.group(1).split(',')]
        return []
    return [main]


def read_labels(filepath, sheet_name, label_col=12, diff_col=4, error_type_col=10, turn_col=11):
    """엑셀 라벨링 시트에서 데이터 읽기"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    items = []
    for row in range(2, ws.max_row + 1):
        no = ws.cell(row=row, column=1).value
        if no is None:
            break
        items.append({
            'no': no,
            'db_id': ws.cell(row=row, column=3).value,
            'difficulty': ws.cell(row=row, column=diff_col).value,
            'error_type': ws.cell(row=row, column=error_type_col).value,
            'turns_used': ws.cell(row=row, column=turn_col).value,
            'label': ws.cell(row=row, column=label_col).value,
            'memo': ws.cell(row=row, column=13).value,
        })
    return items


def main():
    parser = argparse.ArgumentParser(description="유형별 교정률 분석")
    parser.add_argument("--healed", required=True, help="교정 성공 라벨링 시트")
    parser.add_argument("--failed", required=True, help="교정 실패 라벨링 시트")
    args = parser.parse_args()

    healed_items = read_labels(args.healed, '교정 성공 라벨링')
    failed_items = read_labels(args.failed, '오류 라벨링')

    print(f"📂 교정 성공: {len(healed_items)}건")
    print(f"📂 교정 실패: {len(failed_items)}건")
    print(f"📂 전체: {len(healed_items) + len(failed_items)}건")

    # ============================================================
    # 1. 유형별 교정률 (단독 + COMPLEX 내부 합산)
    # ============================================================
    heal_types = defaultdict(int)
    fail_types = defaultdict(int)

    for item in healed_items:
        for t in extract_types(item['label']):
            heal_types[t] += 1

    for item in failed_items:
        for t in extract_types(item['label']):
            fail_types[t] += 1

    all_type_names = sorted(set(list(heal_types.keys()) + list(fail_types.keys())))

    print(f"\n{'='*70}")
    print(f"{'유형':<20} {'교정성공':>8} {'교정실패':>8} {'전체':>8} {'교정률':>10}")
    print(f"{'='*70}")

    for t in all_type_names:
        h = heal_types[t]
        f = fail_types[t]
        total = h + f
        rate = h / total * 100 if total > 0 else 0
        print(f"{t:<20} {h:>8} {f:>8} {total:>8} {rate:>9.1f}%")

    print(f"{'='*70}")

    # ============================================================
    # 2. 메인 라벨 기준 분포 (COMPLEX는 별도)
    # ============================================================
    print(f"\n{'='*70}")
    print("메인 라벨 분포")
    print(f"{'='*70}")

    heal_main = defaultdict(int)
    fail_main = defaultdict(int)

    for item in healed_items:
        label = str(item['label'] or '').strip()
        main = label.split('(')[0].strip() if '(' in label else label
        heal_main[main] += 1

    for item in failed_items:
        label = str(item['label'] or '').strip()
        main = label.split('(')[0].strip() if '(' in label else label
        fail_main[main] += 1

    all_mains = sorted(set(list(heal_main.keys()) + list(fail_main.keys())))

    for m in all_mains:
        h = heal_main[m]
        f = fail_main[m]
        total = h + f
        rate = h / total * 100 if total > 0 else 0
        print(f"  {m:<20} {h:>5} / {total:<5} ({rate:.1f}%)")

    # ============================================================
    # 3. 난이도별 교정률
    # ============================================================
    print(f"\n{'='*70}")
    print("난이도별 교정률")
    print(f"{'='*70}")

    for diff in ['simple', 'moderate', 'challenging']:
        h = sum(1 for item in healed_items if item['difficulty'] == diff)
        f = sum(1 for item in failed_items if item['difficulty'] == diff)
        total = h + f
        rate = h / total * 100 if total > 0 else 0
        print(f"  {diff:<15} {h:>5} / {total:<5} ({rate:.1f}%)")

    # ============================================================
    # 4. DB별 교정률
    # ============================================================
    print(f"\n{'='*70}")
    print("DB별 교정률")
    print(f"{'='*70}")

    heal_db = defaultdict(int)
    fail_db = defaultdict(int)

    for item in healed_items:
        heal_db[item['db_id']] += 1
    for item in failed_items:
        fail_db[item['db_id']] += 1

    all_dbs = sorted(set(list(heal_db.keys()) + list(fail_db.keys())))

    for db in all_dbs:
        h = heal_db[db]
        f = fail_db[db]
        total = h + f
        rate = h / total * 100 if total > 0 else 0
        print(f"  {db:<30} {h:>5} / {total:<5} ({rate:.1f}%)")

    # ============================================================
    # 5. 턴별 교정 분포 (성공 건만)
    # ============================================================
    print(f"\n{'='*70}")
    print("턴별 교정 성공 분포")
    print(f"{'='*70}")

    turn_dist = defaultdict(int)
    for item in healed_items:
        turn_dist[item['turns_used']] += 1

    for t in sorted(turn_dist.keys()):
        print(f"  {t}턴: {turn_dist[t]}건")

    # ============================================================
    # 6. 유형 × 난이도 교차 분석 (교정률)
    # ============================================================
    print(f"\n{'='*70}")
    print("유형 × 난이도 교차 교정률")
    print(f"{'='*70}")
    print(f"{'유형':<20} {'simple':>12} {'moderate':>12} {'challenging':>12}")
    print(f"{'-'*56}")

    for t in all_type_names:
        parts = []
        for diff in ['simple', 'moderate', 'challenging']:
            h = sum(1 for item in healed_items if diff == item['difficulty'] and t in extract_types(item['label']))
            f = sum(1 for item in failed_items if diff == item['difficulty'] and t in extract_types(item['label']))
            total = h + f
            if total > 0:
                parts.append(f"{h}/{total}({h/total*100:.0f}%)")
            else:
                parts.append("-")
        print(f"{t:<20} {parts[0]:>12} {parts[1]:>12} {parts[2]:>12}")


if __name__ == "__main__":
    main()